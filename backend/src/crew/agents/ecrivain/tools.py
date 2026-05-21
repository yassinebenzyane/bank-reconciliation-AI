"""Outil déterministe d'écriture dans la matrice Excel ECO Steering.

Règle absolue : load_workbook(data_only=False, keep_vba=True).
Ne jamais passer data_only=True — cela efface les formules SUMIF du Budget de tréso.
"""
import json
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl
from crewai.tools import BaseTool
from pydantic import BaseModel, PrivateAttr

log = logging.getLogger(__name__)

# Colonnes de l'onglet Base CLient (1-indexé, basé sur l'exploration du fichier réel)
_COL_N_FACTURE = 4   # D — N° FACTURE
_COL_STATUT    = 14  # N — STATUT (a payer / paye)
_COL_DATE_P    = 15  # O — DATE PAIEMENT
_COL_MOIS_P    = 16  # P — MOIS PAIEMENT
_COL_MOYEN     = 17  # Q — MOYEN DE PAIEMENT
_COL_REF       = 18  # R — Référence Paiement

_SHEET = "Base CLient"


class WriteInput(BaseModel):
    matches_json: str   # JSON array des résultats du rapprochement (verificateur)
    xlsx_path:    str   # chemin de la matrice source
    output_path:  str   # chemin de la matrice mise à jour


class UpdateBaseClientTool(BaseTool):
    """Met à jour la matrice Excel ECO Steering avec les résultats du rapprochement validé."""

    name: str        = "update_base_client"
    _result: dict    = PrivateAttr(default_factory=dict)
    description: str = (
        "Met à jour l'onglet Base CLient de la matrice Excel ECO Steering "
        "avec les paiements rapprochés et validés par le comptable. "
        "Pour chaque facture matchée, renseigne : STATUT (payé), DATE PAIEMENT, "
        "MOIS PAIEMENT, MOYEN DE PAIEMENT (virement), Référence Paiement. "
        "Ne modifie jamais les factures non rapprochées ni les colonnes hors périmètre."
    )
    args_schema: type[BaseModel] = WriteInput

    def _run(self, matches_json: str, xlsx_path: str, output_path: str) -> str:
        matches = json.loads(matches_json)
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Construire un index facture_id → données de paiement depuis les matches
        paiements: dict[str, dict] = {}
        for m in matches:
            if m.get("statut") != "match_exact":
                continue
            txn       = m.get("transaction", {})
            date_op   = _parse_date(txn.get("date_operation"))
            reference = str(txn.get("reference") or "").strip()
            mois_p    = date_op.replace(day=1) if date_op else None
            for fid in m.get("factures_ids", []):
                paiements[str(fid).strip()] = {"date_p": date_op, "mois_p": mois_p, "ref": reference}

        if not paiements:
            return json.dumps({"updated": 0, "message": "Aucun match exact à écrire."})

        # Passe 1 — data_only=True : cache les vraies valeurs J/K (shared-formulas)
        wb_ro = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws_ro = wb_ro[_SHEET]
        try:
            if ws_ro._cells:
                ws_ro._max_row = max(r for r, _ in ws_ro._cells)
        except (AttributeError, ValueError, TypeError):
            pass
        jk_cache: dict[int, tuple] = {}  # clé = row_num (unique même si n_facture dupliqué)
        for rn in range(4, ws_ro.max_row + 1):
            nf = str(ws_ro.cell(row=rn, column=_COL_N_FACTURE).value or "").strip()
            if nf:
                jk_cache[rn] = (
                    ws_ro.cell(row=rn, column=10).value,
                    ws_ro.cell(row=rn, column=11).value,
                )
        wb_ro.close()

        # Passe 2 — data_only=False : préserve les formules Budget de tréso
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        ws = wb[_SHEET]
        try:
            if ws._cells:
                ws._max_row    = max(r for r, _ in ws._cells)
                ws._max_column = max(c for _, c in ws._cells)
        except (AttributeError, ValueError, TypeError):
            pass

        updated = 0
        skipped = []

        for row_num in range(4, ws.max_row + 1):
            n_fact = str(ws.cell(row=row_num, column=_COL_N_FACTURE).value or "").strip()
            if n_fact not in paiements:
                continue
            statut_actuel = str(ws.cell(row=row_num, column=_COL_STATUT).value or "").strip().lower()
            if statut_actuel in {"paye", "payé"}:
                skipped.append(n_fact + " (déjà payé)")
                continue

            p = paiements[n_fact]
            ws.cell(row=row_num, column=_COL_STATUT).value = "payé"
            ws.cell(row=row_num, column=_COL_DATE_P ).value = p["date_p"]
            ws.cell(row=row_num, column=_COL_MOIS_P ).value = p["mois_p"]
            ws.cell(row=row_num, column=_COL_MOYEN  ).value = "virement"
            ws.cell(row=row_num, column=_COL_REF    ).value = p["ref"]

            # Réinjecter J et K depuis le cache (clé = row_num, pas n_facture)
            if row_num in jk_cache:
                val_j, val_k = jk_cache[row_num]
                if val_j is not None:
                    ws.cell(row=row_num, column=10).value = val_j
                if val_k is not None:
                    ws.cell(row=row_num, column=11).value = val_k

            updated += 1
            log.info("Écrivain : R%d %s → payé %s", row_num, n_fact, p["date_p"])

        wb.save(output_path)
        wb.close()

        msg = f"{updated} ligne(s) mises à jour dans Base CLient."
        if skipped:
            msg += f" Ignorées (déjà payées) : {skipped}."
        log.info("Écrivain : %s", msg)
        self._result = {"updated": updated, "skipped": skipped, "output": output_path, "message": msg}
        return json.dumps(self._result)

    @property
    def result(self) -> dict:
        return self._result


def _parse_date(raw) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return datetime(raw.year, raw.month, raw.day) if isinstance(raw, date) else raw
    try:
        return datetime.fromisoformat(str(raw)[:10])
    except ValueError:
        return None
