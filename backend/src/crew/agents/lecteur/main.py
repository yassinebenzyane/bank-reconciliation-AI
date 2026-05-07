"""Agent Lecteur — orchestrateur parsers Python + validation."""
from __future__ import annotations
import logging
import time
from datetime import date
from pathlib import Path

import openpyxl

from .config import (
    TABLE_CONFIGS, MATRIX_CONFIGS, PRET_CONFIG,
    BANK_CSV_CONFIG, LABEL_CACHE_FILE, CACHE_DIR,
)
from .parsers.table_parser       import TableParser
from .parsers.matrix_parser      import MatrixParser
from .parsers.key_value_parser   import KeyValueParser
from .parsers.bank_statement_parser import BankStatementParser
from .validators                 import validate_all
from .semantic.llm_client        import LLMClient
from .semantic.label_normalizer  import LabelNormalizer
from .semantic.field_decoder     import FieldDecoder

log = logging.getLogger(__name__)


def run(
    xlsx_path: str,
    csv_path:  str | None = None,
    target_month: date | None = None,
    use_semantic: bool = True,
    detect_drift: bool = False,
) -> dict:
    t0 = time.time()
    result: dict = {}
    llm_calls = 0

    log.info("Ouverture de la matrice Excel...")
    # data_only=True requis pour lire les valeurs calculées (Prêt B11/B12).
    # Ne pas passer read_only=True : restreint iter_rows au tag <dimension> XML
    # qui est souvent sous-estimé (ex: déclare 1813 lignes au lieu de 2329).
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheets = wb.sheetnames

    for sheet_name, cfg in TABLE_CONFIGS.items():
        if sheet_name not in sheets:
            log.warning("Onglet absent : %s", sheet_name)
            continue
        log.info("Parsing %s...", sheet_name)
        pr = TableParser(wb[sheet_name], sheet_name, cfg).parse()
        result[sheet_name] = pr.data
        _log_result(pr)

    months = [target_month] if target_month else None
    for sheet_name, cfg in MATRIX_CONFIGS.items():
        if sheet_name not in sheets:
            continue
        log.info("Parsing %s...", sheet_name)
        pr = MatrixParser(wb[sheet_name], sheet_name, cfg).parse(target_months=months)
        result[sheet_name] = pr.data
        _log_result(pr)

    if "Prêt" in sheets:
        log.info("Parsing Prêt...")
        pr = KeyValueParser(wb["Prêt"], "Prêt", PRET_CONFIG).parse()
        result["Prêt"] = pr.data[0] if pr.data else {}

    wb.close()

    if csv_path:
        log.info("Parsing relevé bancaire CSV...")
        cfg = BANK_CSV_CONFIG["BPVF"]
        pr  = BankStatementParser(str(csv_path), cfg).parse()
        result["releve_bancaire"] = pr.data
        _log_result(pr)

    if use_semantic:
        llm = LLMClient(cache_dir=CACHE_DIR)

        budget_data = result.get("Budget de tréso", [])
        if budget_data:
            log.info("Normalisation des labels Budget de tréso (LLM + cache)...")
            key_rows = MATRIX_CONFIGS["Budget de tréso"]["key_rows"]
            labels = [(rn, field) for rn, field in key_rows.items()]
            try:
                normalizer = LabelNormalizer(LABEL_CACHE_FILE, llm)
                normalized = normalizer.normalize(labels)
                result["_budget_labels"] = normalized
                llm_calls += 1
            except Exception as e:
                log.warning("Label normalizer échoué : %s", e)

        ndf_data = result.get("NdF", [])
        if ndf_data:
            log.info("Décodage champ libre NdF.salarie (LLM batch)...")
            texts = [r.get("salarie") or "" for r in ndf_data]
            try:
                decoder  = FieldDecoder(llm)
                decoded  = decoder.decode_batch(texts)
                llm_calls += 1
                for row, dec in zip(ndf_data, decoded):
                    row["salarie_decoded"] = dec
            except Exception as e:
                log.warning("FieldDecoder échoué : %s", e)

    log.info("Validation...")
    validation_errors = validate_all(result)
    result["_validation"] = validation_errors
    if validation_errors:
        for sheet, errs in validation_errors.items():
            for e in errs:
                log.warning("[VALIDATION] %s : %s", sheet, e)

    result["_meta"] = {
        "xlsx": str(xlsx_path),
        "csv":  str(csv_path) if csv_path else None,
        "duration_s":  round(time.time() - t0, 2),
        "llm_calls":   llm_calls,
        "sheets_parsed": [s for s in TABLE_CONFIGS if s in result]
                        + [s for s in MATRIX_CONFIGS if s in result],
        "validation_ok": not validation_errors,
    }

    log.info(
        "Agent Lecteur terminé en %.1fs | %d appels LLM | validation=%s",
        result["_meta"]["duration_s"],
        llm_calls,
        "OK" if not validation_errors else f"{sum(len(e) for e in validation_errors.values())} erreurs",
    )
    return result


def _log_result(pr) -> None:
    log.info(
        "  %s : %d/%d lignes extraites, %d warnings",
        pr.sheet, pr.row_count_extracted, pr.row_count_source, len(pr.warnings),
    )
    for w in pr.warnings:
        log.warning("  [%s] %s", pr.sheet, w)
